#! python3
# excel_to_csv_converter.py — An exercise in manipulating CSV files.
# For more information, see project_details.txt.

import csv
import logging
import os
import openpyxl
from openpyxl.utils import get_column_letter

logging.basicConfig(
    level=logging.DEBUG,
    filename="logging.txt",
    format="%(asctime)s -  %(levelname)s -  %(message)s",
)
# logging.disable(logging.CRITICAL)  # Note out to enable logging.

for excelFile in os.listdir("."):
    # Skip non-xlsx files, load the workbook object.
    if excelFile.endswith(".xlsx"):
        wb = openpyxl.load_workbook(excelFile)

        for sheetName in wb.sheetnames:
            # Loop through every sheet in the workbook.
            sheet = wb[sheetName]

            # Create the CSV filename from the Excel filename and sheet title.
            csv_filename = f"{excelFile[:-5]}_{sheetName}.csv"
            # logging.info(csv_filename)

            # Create the csv.writer object for this CSV file.
            csv_file = open(csv_filename, "w", newline="")
            csv_writer = csv.writer(csv_file)
            # Loop through every row in the sheet.
            for row_num in range(1, sheet.max_row + 1):
                row_data = []  # append each cell to this list
                # Loop through each cell in the row.
                for col_num in range(1, sheet.max_column + 1):
                    # Append each cell's data to rowData.
                    column_letter = get_column_letter(col_num)
                    row_data.append(sheet[f"{column_letter}{row_num}"].value)
                logging.info(row_data)

                # Write the rowData list to the CSV file.
                csv_writer.writerow(row_data)

            csv_file.close()
